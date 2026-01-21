import os
import io
from datetime import date, datetime

from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
)
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_JUSTIFY, TA_CENTER

from PySide6.QtWidgets import QFileDialog, QMessageBox

from paths import ICON_DIR
from models.institucion_model import InstitucionModel
from models.secciones_model import SeccionesModel
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from utils.edad import calcular_edad
from utils.dialogs import crear_msgbox

# Estilos globales
styles = getSampleStyleSheet()

justificado = ParagraphStyle(
    name="Justificado",
    parent=styles["Normal"],
    alignment=TA_JUSTIFY,
    leading=16
)
centrado = ParagraphStyle(
    name="Centrado",
    parent=styles["Normal"],
    alignment=TA_CENTER,
    leading=16
)

page_width, page_height = letter


# UTILIDADES DE VALIDACIÓN Y CONVERSIÓN

def sanitizar_nombre_archivo(nombre: str) -> str:
    """Limpia caracteres inválidos de nombres de archivo."""
    caracteres_invalidos = r'<>:"/\|?*'
    for char in caracteres_invalidos:
        nombre = nombre.replace(char, '_')
    return nombre.strip()


def crear_carpeta_segura(ruta: str) -> tuple[bool, str]:
    """Crea una carpeta verificando permisos de escritura."""
    try:
        os.makedirs(ruta, exist_ok=True)
        
        # Verificar que sea escribible
        test_file = os.path.join(ruta, '.test_write')
        try:
            with open(test_file, 'w') as f:
                f.write('test')
            os.remove(test_file)
            return True, "Carpeta creada correctamente"
        except Exception as e:
            return False, f"La carpeta no tiene permisos de escritura: {e}"
            
    except Exception as e:
        return False, f"No se pudo crear la carpeta: {e}"


def convertir_fecha_string(fecha) -> str:
    """Convierte diferentes formatos de fecha a string DD/MM/YYYY."""
    if fecha is None:
        return "N/A"
    
    if isinstance(fecha, (date, datetime)):
        return fecha.strftime("%d/%m/%Y")
    
    # Si ya es string, intentar parsearlo para validar
    if isinstance(fecha, str):
        try:
            fecha_obj = datetime.strptime(fecha, "%d/%m/%Y")
            return fecha_obj.strftime("%d/%m/%Y")
        except:
            return str(fecha)
    
    return str(fecha)


def extraer_año_escolar(año_escolar: dict) -> tuple[int, int]:
    """Extrae años de inicio y fin del diccionario año_escolar."""
    if not año_escolar:
        año_actual = datetime.now().year
        return año_actual, año_actual + 1
    
    anio_inicio = año_escolar.get('anio_inicio')
    
    if isinstance(anio_inicio, (date, datetime)):
        año_inicio = anio_inicio.year
    elif isinstance(anio_inicio, str):
        try:
            año_inicio = int(anio_inicio.split('-')[0])
        except:
            año_inicio = datetime.now().year
    else:
        año_inicio = int(anio_inicio) if anio_inicio else datetime.now().year
    
    año_fin = año_inicio + 1
    return año_inicio, año_fin


def normalizar_cedula(cedula: str, es_estudiante: bool = False) -> str:
    """Normaliza formato de cédula, agregando prefijo V- o CE- si no tiene."""
    if not cedula:
        return "N/A"
    
    cedula = str(cedula).strip().upper()
    
    # Si ya tiene prefijo (V-, E-, J-, G-, CE-), mantenerlo
    if len(cedula) > 2:
        if cedula[:2] == 'CE' and cedula[2] == '-':
            return cedula
        if cedula[0] in ['V', 'E', 'J', 'G'] and cedula[1] == '-':
            return cedula
    
    # Si no tiene prefijo, agregar según el tipo
    prefijo = "CE-" if es_estudiante else "V-"
    return f"{prefijo}{cedula}"


def validar_datos_exportacion(datos: dict, campos_requeridos: list) -> tuple[bool, str]:
    """Valida que los datos tengan los campos requeridos para exportación."""
    if not datos:
        return False, "No se proporcionaron datos"
    
    campos_faltantes = []
    for campo in campos_requeridos:
        if campo not in datos or datos[campo] is None or str(datos[campo]).strip() == "":
            campos_faltantes.append(campo)
    
    if campos_faltantes:
        return False, f"Faltan campos requeridos: {', '.join(campos_faltantes)}"
    
    return True, "Datos válidos"


# FUNCIONES DE ENCABEZADO Y PIE DE PÁGINA

def draw_centered(canvas, text, y, font="Helvetica", size=12):
    """Dibuja texto centrado en el canvas."""
    canvas.setFont(font, size)
    text_width = canvas.stringWidth(text, font, size)
    x_center = (page_width - text_width) / 2
    canvas.drawString(x_center, y, text)


def encabezado(canvas, doc, institucion_id=1):
    """
    Dibuja el encabezado con logo e institución en cada página del PDF.
    """
    logo_path = os.path.join(ICON_DIR, "logo_escuela_fondo.png")
    
    # Verificar que el logo existe
    if not os.path.exists(logo_path):
        print(f"Advertencia: Logo no encontrado en {logo_path}")
        # Continuar sin logo
    else:
        # Dimensiones del logo
        logo_width = 65
        logo_height = 60

        # Coordenadas
        x_center = (page_width - logo_width) / 2
        y_position = page_height - 180

        try:
            canvas.drawImage(
                logo_path,
                x=x_center,
                y=y_position,
                width=logo_width,
                height=logo_height,
                preserveAspectRatio=True
            )
        except Exception as e:
            print(f"Error dibujando logo: {e}")

    # Obtener datos de institución
    institucion = InstitucionModel.obtener_por_id(institucion_id)

    if institucion:
        nombre = str(institucion.get("nombre", "")).upper()
        codigo = str(institucion.get("codigo_dea", ""))

        draw_centered(canvas, "REPÚBLICA BOLIVARIANA DE VENEZUELA", page_height - 50, "Helvetica", 10)
        draw_centered(canvas, "MINISTERIO DEL PODER POPULAR PARA LA EDUCACIÓN", page_height - 65, "Helvetica", 10)
        draw_centered(canvas, nombre, page_height - 80, "Helvetica", 10)
        draw_centered(canvas, f"CÓDIGO DEA: {codigo}", page_height - 95, "Helvetica", 10)
        draw_centered(canvas, "PUERTO LA CRUZ, EDO. ANZOÁTEGUI", page_height - 110, "Helvetica", 10)
    else:
        canvas.setFont("Helvetica-Bold", 12)
        fallback = "Institución no encontrada"
        text_width = canvas.stringWidth(fallback, "Helvetica-Bold", 12)
        x_center = (page_width - text_width) / 2
        canvas.drawString(x_center, page_height - 50, fallback)


def pie_pagina(canvas, doc, institucion_id=1):
    """
    Dibuja el pie de página con datos de la institución.
    """
    institucion = InstitucionModel.obtener_por_id(institucion_id)

    font_name = "Helvetica"
    font_size = 10
    canvas.setFont(font_name, font_size)

    if institucion:
        direccion = str(institucion.get("direccion", "")).upper()
        telefono = str(institucion.get("telefono", ""))
        correo = str(institucion.get("correo", "")).upper()

        line1 = f"DIRECCIÓN: {direccion}" if direccion else ""
        line2 = f"TELÉFONO: {telefono} | CORREO: {correo}" if telefono or correo else ""
    else:
        line1 = ""
        line2 = ""

    for i, line in enumerate([line1, line2]):
        if line.strip():
            text_width = canvas.stringWidth(line, font_name, font_size)
            x_center = (page_width - text_width) / 2
            y_pos = 45 - i * (font_size + 2)
            canvas.drawString(x_center, y_pos, line)


def encabezado_y_pie(canvas, doc):
    """
    Combina encabezado y pie de página.
    """
    encabezado(canvas, doc)
    pie_pagina(canvas, doc)


def encabezado_prosecucion(canvas, doc):
    """
    Encabezado personalizado para constancia de prosecución.
    """
    # 1. Logo esquinado del ministerio (arriba a la izquierda)
    logo_ministerio_path = os.path.join(ICON_DIR, "ministerio.png")
    
    if os.path.exists(logo_ministerio_path):
        logo_min_width = 200
        logo_min_height = 120
        x_esquina = 40  # Margen izquierdo
        y_esquina = page_height - 100  # Posición superior

        try:
            canvas.drawImage(
                logo_ministerio_path,
                x_esquina,
                y_esquina,
                width=logo_min_width,
                height=logo_min_height,
                preserveAspectRatio=True,
                mask='auto'
            )
        except Exception as e:
            print(f"Error cargando logo del ministerio: {e}")
    
    # 2. Escudo centrado de Venezuela
    logo_venezuela_path = os.path.join(ICON_DIR, "Escudo_Venezuela.png")
        
    if os.path.exists(logo_venezuela_path):
        logo_ven_width = 70
        logo_ven_height = 85

        # Coordenadas centradas, debajo del logo del ministerio
        x_center = (page_width - logo_ven_width) / 2
        y_position = page_height - 170

        try:
            canvas.drawImage(
                logo_venezuela_path,
                x_center,
                y_position,
                width=logo_ven_width,
                height=logo_ven_height,
                preserveAspectRatio=True,
                mask='auto'
            )
        except Exception as e:
            print(f"Error cargando logo de Venezuela: {e}")


# FORMATOS ESTUDIANTES

def generar_constancia_estudios(estudiante: dict, institucion: dict) -> str:
    """Genera constancia de estudios en PDF para un estudiante."""
    # Validar datos requeridos
    campos_est = ["Nombres", "Apellidos", "Cédula", "Grado", "Sección"]
    valido, mensaje = validar_datos_exportacion(estudiante, campos_est)
    if not valido:
        raise ValueError(f"Datos de estudiante incompletos: {mensaje}")
    
    campos_inst = ["director", "director_ci", "nombre"]
    valido, mensaje = validar_datos_exportacion(institucion, campos_inst)
    if not valido:
        raise ValueError(f"Datos de institución incompletos: {mensaje}")
    
    # Normalizar datos
    estudiante["Nombres"] = str(estudiante["Nombres"]).strip().upper()
    estudiante["Apellidos"] = str(estudiante["Apellidos"]).strip().upper()
    cedula_normalizada = normalizar_cedula(estudiante["Cédula"], es_estudiante=True)
    
    # Crear carpeta
    carpeta = os.path.join(os.getcwd(), "exportados", "Constancias de estudios")
    ok, msg = crear_carpeta_segura(carpeta)
    if not ok:
        raise IOError(msg)

    # Nombre de archivo sanitizado
    nombre_base = sanitizar_nombre_archivo(f"Constancia_{estudiante['Cédula']}")
    nombre_archivo = os.path.join(carpeta, f"{nombre_base}.pdf")

    try:
        doc = SimpleDocTemplate(
            nombre_archivo,
            pagesize=letter,
            leftMargin=80,
            rightMargin=80,
            topMargin=180,
            bottomMargin=50
        )

        story = []

        # Título
        story.append(Paragraph("CONSTANCIA DE ESTUDIOS", styles["Title"]))
        story.append(Spacer(1, 16))

        # Texto principal
        director_ci = normalizar_cedula(institucion['director_ci'])
        texto = (
            f"El suscrito, Director <b>PROF. {institucion['director'].upper()}</b>, portador de la Cédula de Identidad "
            f"<b>{director_ci}</b>, de {institucion['nombre']}, hace constar que el(la) estudiante "
            f"<b>{estudiante['Apellidos']} {estudiante['Nombres']}</b>, "
            f"portador de la cédula escolar <b>{cedula_normalizada}</b>, cursa actualmente el "
            f"<b>{estudiante['Grado']} grado Sección '{estudiante['Sección']}'</b> de Educación Primaria en esta institución.<br/><br/>"
        )
        story.append(Paragraph(texto, justificado))
        story.append(Spacer(1, 40))

        # Fecha
        fecha_hoy = date.today().strftime("%d/%m/%Y")
        texto_fecha = (
            "Constancia que se expide a petición de la parte interesada en la Ciudad de Puerto La Cruz, "
            f"a la fecha {fecha_hoy}."
        )
        story.append(Paragraph(texto_fecha, justificado))
        story.append(Spacer(1, 100))

        # Firma
        firma = f"________________________<br/>Prof. {institucion['director']}"
        story.append(Paragraph(firma, centrado))
        story.append(Spacer(1, 3))
        story.append(Paragraph(f"C.I. {director_ci}", centrado))
        story.append(Spacer(1, 3))
        story.append(Paragraph("Director", centrado))

        # Construir PDF
        doc.build(story, onFirstPage=encabezado_y_pie, onLaterPages=encabezado_y_pie)
        return nombre_archivo
        
    except Exception as e:
        raise IOError(f"Error generando PDF: {e}")


def generar_buena_conducta(estudiante: dict, institucion: dict, año_escolar: dict) -> str:
    """Genera constancia de buena conducta en PDF para un estudiante."""
    # Validar datos
    campos_est = ["Nombres", "Apellidos", "Cédula", "Grado", "Sección"]
    valido, mensaje = validar_datos_exportacion(estudiante, campos_est)
    if not valido:
        raise ValueError(f"Datos de estudiante incompletos: {mensaje}")
    
    campos_inst = ["director", "director_ci", "nombre"]
    valido, mensaje = validar_datos_exportacion(institucion, campos_inst)
    if not valido:
        raise ValueError(f"Datos de institución incompletos: {mensaje}")
    
    # Extraer años
    año_inicio, año_fin = extraer_año_escolar(año_escolar)
    
    # Normalizar datos
    estudiante["Nombres"] = str(estudiante["Nombres"]).strip().upper()
    estudiante["Apellidos"] = str(estudiante["Apellidos"]).strip().upper()
    cedula_normalizada = normalizar_cedula(estudiante["Cédula"], es_estudiante=True)

    # Crear carpeta
    carpeta = os.path.join(os.getcwd(), "exportados", "Constancias de buena conducta")
    ok, msg = crear_carpeta_segura(carpeta)
    if not ok:
        raise IOError(msg)

    nombre_base = sanitizar_nombre_archivo(f"Constancia_buena_conducta_{estudiante['Cédula']}")
    nombre_archivo = os.path.join(carpeta, f"{nombre_base}.pdf")

    try:
        doc = SimpleDocTemplate(
            nombre_archivo,
            pagesize=letter,
            leftMargin=80,
            rightMargin=80,
            topMargin=180,
            bottomMargin=50
        )

        story = []

        # Título
        story.append(Paragraph("CONSTANCIA DE BUENA CONDUCTA", styles["Title"]))
        story.append(Spacer(1, 16))

        # Texto principal
        director_ci = normalizar_cedula(institucion['director_ci'])
        texto = (
            f"El suscrito, Director <b>PROF. {institucion['director'].upper()}</b>, portador de la Cédula de Identidad "
            f"<b>{director_ci}</b>, de {institucion['nombre']}, que funciona en Puerto La Cruz, "
            f"hace constar que el alumno(a): <b>{estudiante['Apellidos']} {estudiante['Nombres']}</b>, "
            f"portador de la cédula de identidad <b>{cedula_normalizada}</b>, estudiante del "
            f"<b>{estudiante['Grado']} Grado Sección '{estudiante['Sección']}'</b> "
            f"de Educación Primaria durante el Año Escolar {año_inicio}-{año_fin}, mantuvo una "
            f"<b>Buena Conducta</b> durante su permanencia en esta institución educativa.<br/><br/>"
        )
        story.append(Paragraph(texto, justificado))
        story.append(Spacer(1, 40))

        # Fecha
        fecha_hoy = date.today().strftime("%d/%m/%Y")
        texto_fecha = (
            "Constancia que se expide a petición de la parte interesada en la Ciudad de Puerto La Cruz, "
            f"a la fecha {fecha_hoy}."
        )
        story.append(Paragraph(texto_fecha, justificado))
        story.append(Spacer(1, 100))

        # Firma
        firma = f"________________________<br/>Prof. {institucion['director']}"
        story.append(Paragraph(firma, centrado))
        story.append(Spacer(1, 3))
        story.append(Paragraph(f"C.I. {director_ci}", centrado))
        story.append(Spacer(1, 3))
        story.append(Paragraph("Director", centrado))

        # Construir PDF
        doc.build(story, onFirstPage=encabezado_y_pie, onLaterPages=encabezado_y_pie)
        return nombre_archivo
        
    except Exception as e:
        raise IOError(f"Error generando PDF: {e}")


def generar_constancia_inscripcion(estudiante: dict, institucion: dict) -> str:
    """Genera constancia de inscripción en PDF para un estudiante."""
    # Validar datos
    campos_est = ["Nombres", "Apellidos", "Cédula", "Grado", "Ciudad", "Fecha Nac.", "Fecha Ingreso"]
    valido, mensaje = validar_datos_exportacion(estudiante, campos_est)
    if not valido:
        raise ValueError(f"Datos de estudiante incompletos: {mensaje}")
    
    campos_inst = ["director", "director_ci"]
    valido, mensaje = validar_datos_exportacion(institucion, campos_inst)
    if not valido:
        raise ValueError(f"Datos de institución incompletos: {mensaje}")
    
    # Normalizar datos
    estudiante["Nombres"] = str(estudiante["Nombres"]).strip().upper()
    estudiante["Apellidos"] = str(estudiante["Apellidos"]).strip().upper()
    cedula_normalizada = normalizar_cedula(estudiante["Cédula"], es_estudiante=True)
    
    # Convertir fechas
    fecha_nac_str = convertir_fecha_string(estudiante['Fecha Nac.'])
    fecha_ingreso_str = convertir_fecha_string(estudiante['Fecha Ingreso'])
    
    # Calcular edad
    try:
        fecha_nac = estudiante['Fecha Nac.']
        if isinstance(fecha_nac, (date, datetime)):
            edad = calcular_edad(fecha_nac)
        else:
            try:
                fecha_obj = datetime.strptime(fecha_nac_str, "%d/%m/%Y").date()
                edad = calcular_edad(fecha_obj)
            except:
                edad = "N/A"
    except:
        edad = "N/A"

    # Crear carpeta
    carpeta = os.path.join(os.getcwd(), "exportados", "Constancias de inscripcion")
    ok, msg = crear_carpeta_segura(carpeta)
    if not ok:
        raise IOError(msg)

    nombre_base = sanitizar_nombre_archivo(f"Constancia_inscripcion_{estudiante['Cédula']}")
    nombre_archivo = os.path.join(carpeta, f"{nombre_base}.pdf")

    try:
        doc = SimpleDocTemplate(
            nombre_archivo,
            pagesize=letter,
            leftMargin=80,
            rightMargin=80,
            topMargin=180,
            bottomMargin=50
        )

        story = []

        # Título
        story.append(Paragraph("CONSTANCIA DE INSCRIPCIÓN", styles["Title"]))
        story.append(Spacer(1, 16))

        # Texto principal
        texto = (
            f"La Dirección del plantel hace constar mediante la presente, que el Alumno(a): "
            f"<b>{estudiante['Apellidos']} {estudiante['Nombres']}</b>, nacido en {estudiante['Ciudad']} "
            f"en fecha {fecha_nac_str}, de {edad} años de edad, fué inscrito "
            f"en esta institución el día {fecha_ingreso_str} para cursar el <b>{estudiante['Grado']} Grado</b> "
            f"de Educación Primaria."
        )
        story.append(Paragraph(texto, justificado))
        story.append(Spacer(1, 40))

        # Fecha
        fecha_hoy = date.today().strftime("%d/%m/%Y")
        texto_fecha = (
            "Constancia que se expide a petición de la parte interesada en la Ciudad de Puerto La Cruz, "
            f"a la fecha {fecha_hoy}."
        )
        story.append(Paragraph(texto_fecha, justificado))
        story.append(Spacer(1, 100))

        # Firma
        director_ci = normalizar_cedula(institucion['director_ci'])
        firma = f"________________________<br/>Prof. {institucion['director']}"
        story.append(Paragraph(firma, centrado))
        story.append(Spacer(1, 3))
        story.append(Paragraph(f"C.I. {director_ci}", centrado))
        story.append(Spacer(1, 3))
        story.append(Paragraph("Director", centrado))

        # Construir PDF
        doc.build(story, onFirstPage=encabezado_y_pie, onLaterPages=encabezado_y_pie)
        return nombre_archivo
        
    except Exception as e:
        raise IOError(f"Error generando PDF: {e}")


def generar_constancia_prosecucion_inicial(estudiante: dict, institucion: dict, año_escolar: dict) -> str:
    """Genera constancia de prosecución de educación inicial a primaria en PDF."""
    # Validar datos
    campos_est = ["Nombres", "Apellidos", "Cédula", "Ciudad", "Fecha Nac."]
    valido, mensaje = validar_datos_exportacion(estudiante, campos_est)
    if not valido:
        raise ValueError(f"Datos de estudiante incompletos: {mensaje}")
    
    campos_inst = ["director", "director_ci", "nombre"]
    valido, mensaje = validar_datos_exportacion(institucion, campos_inst)
    if not valido:
        raise ValueError(f"Datos de institución incompletos: {mensaje}")
    
    # Extraer años
    año_inicio, año_fin = extraer_año_escolar(año_escolar)
    año_anterior = año_inicio - 1
    
    # Normalizar datos
    estudiante["Nombres"] = str(estudiante["Nombres"]).strip().upper()
    estudiante["Apellidos"] = str(estudiante["Apellidos"]).strip().upper()
    cedula_normalizada = normalizar_cedula(estudiante["Cédula"], es_estudiante=True)
    
    # Convertir fecha
    fecha_nac_str = convertir_fecha_string(estudiante['Fecha Nac.'])
    
    # Crear carpeta
    carpeta = os.path.join(os.getcwd(), "exportados", "Constancias de prosecucion inicial")
    ok, msg = crear_carpeta_segura(carpeta)
    if not ok:
        raise IOError(msg)

    nombre_base = sanitizar_nombre_archivo(f"Constancia_prosecucion_inicial_{estudiante['Cédula']}")
    nombre_archivo = os.path.join(carpeta, f"{nombre_base}.pdf")

    try:
        doc = SimpleDocTemplate(
            nombre_archivo,
            pagesize=letter,
            leftMargin=80,
            rightMargin=80,
            topMargin=180,
            bottomMargin=50
        )

        story = []

        # Título
        story.append(Paragraph("CONSTANCIA DE PROSECUSION<br/>EN EL NIVEL DE EDUCACION INICIAL", styles["Title"]))
        story.append(Spacer(1, 16))

        # Texto principal
        director_ci = normalizar_cedula(institucion['director_ci'])
        texto = (
            f"Quien suscribe <b>{institucion['director'].upper()}</b> titular de la Cédula de Identidad "
            f"Nº <b>{director_ci}</b> Director(a) de la Institución Educativa "
            f"<b>{institucion['nombre'].upper()}</b>, ubicada en el Municipio JUAN ANTONIO SOTILLO de la Parroquia "
            f"PUERTO LA CRUZ adscrita al Centro de Desarrollo de la Calidad Educativa Estadal ANZOÁTEGUI. "
            f"Por la presente certifica que el (la) estudiante <b>{estudiante['Apellidos']} {estudiante['Nombres']}</b> "
            f"titular de Cédula Escolar <b>{cedula_normalizada}</b>, nacido (a) en {estudiante['Ciudad']} "
            f"del Estado <b>ANZOÁTEGUI</b> en fecha <b>{fecha_nac_str}</b>, cursó el <b>3er nivel</b> de la etapa de "
            f"<b>Educación Inicial</b> durante el periodo escolar <b>{año_anterior}-{año_inicio}</b>, siendo "
            f"promovido(a) a <b>primer grado de primaria</b>, previo cumplimiento a los requisitos establecidos "
            f"en la normativa legal vigente."
        )
        story.append(Paragraph(texto, justificado))
        story.append(Spacer(1, 40))

        # Fecha de expedición
        fecha_hoy = date.today()
        dia = fecha_hoy.day
        mes_nombre = fecha_hoy.strftime("%B").upper()
        meses = {
            'JANUARY': 'ENERO', 'FEBRUARY': 'FEBRERO', 'MARCH': 'MARZO',
            'APRIL': 'ABRIL', 'MAY': 'MAYO', 'JUNE': 'JUNIO',
            'JULY': 'JULIO', 'AUGUST': 'AGOSTO', 'SEPTEMBER': 'SEPTIEMBRE',
            'OCTOBER': 'OCTUBRE', 'NOVEMBER': 'NOVIEMBRE', 'DECEMBER': 'DICIEMBRE'
        }
        mes_es = meses.get(mes_nombre, mes_nombre)
        año = fecha_hoy.year
        
        texto_fecha = f"Certificado que se expide en PUERTO LA CRUZ, a los {dia} días del mes de {mes_es} de {año}"
        story.append(Paragraph(texto_fecha, justificado))
        story.append(Spacer(1, 30))

        # Tabla de firmas institucionales
        tabla_firmas_data = [
            # Encabezados
            [
                Paragraph("<b>INSTITUCIÓN EDUCATIVA<br/>(PARA VALIDEZ NACIONAL)</b>", centrado),
                Paragraph("<b>CENTRO DE DESARROLLO DE LA CALIDAD<br/>EDUCATIVA ESTADAL<br/>(PARA VALIDEZ INTERNACIONAL)</b>", centrado)
            ],
            # Director(a)
            [
                Paragraph("DIRECTOR(A)", styles["Normal"]),
                Paragraph("DIRECTOR(A)", styles["Normal"])
            ],
            # Nombre y Apellido
            [
                Paragraph(f"Nombre y Apellido: {institucion['director'].upper()}", styles["Normal"]),
                Paragraph("Nombre y Apellido:", styles["Normal"])
            ],
            # Cédula de Identidad
            [
                Paragraph(f"Número de C.I: {director_ci}", styles["Normal"]),
                Paragraph("Número de C.I:", styles["Normal"])
            ],
            # Firma y Sello (espacios vacíos)
            [
                Paragraph("Firma y Sello:<br/><br/><br/>", styles["Normal"]),
                Paragraph("Firma y Sello:<br/><br/><br/>", styles["Normal"])
            ]
        ]

        tabla_firmas = Table(tabla_firmas_data, colWidths=[page_width/2 - 100, page_width/2 - 100])
        tabla_firmas.setStyle(TableStyle([
            # Bordes externos de la tabla
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            
            # Alineación
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),  # Encabezados centrados
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),   # Resto alineado a la izquierda
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),   # Alineación vertical superior
            
            # Padding interno
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        story.append(tabla_firmas)

        # Construir PDF
        doc.build(story, onFirstPage=encabezado_prosecucion, onLaterPages=encabezado_prosecucion)
        return nombre_archivo
        
    except Exception as e:
        raise IOError(f"Error generando PDF: {e}")


# FORMATOS EMPLEADOS

def generar_constancia_trabajo(empleado: dict, institucion: dict) -> str:
    """Genera constancia de trabajo en PDF para un empleado."""
    # Validar datos
    campos_emp = ["Nombres", "Apellidos", "Cédula", "Cargo", "Fecha Ingreso"]
    valido, mensaje = validar_datos_exportacion(empleado, campos_emp)
    if not valido:
        raise ValueError(f"Datos de empleado incompletos: {mensaje}")
    
    campos_inst = ["director", "director_ci", "nombre"]
    valido, mensaje = validar_datos_exportacion(institucion, campos_inst)
    if not valido:
        raise ValueError(f"Datos de institución incompletos: {mensaje}")
    
    # Normalizar datos
    empleado["Nombres"] = str(empleado["Nombres"]).strip().upper()
    empleado["Apellidos"] = str(empleado["Apellidos"]).strip().upper()
    cedula_normalizada = normalizar_cedula(empleado["Cédula"])
    
    # Convertir fecha
    fecha_ingreso_str = convertir_fecha_string(empleado['Fecha Ingreso'])
    
    # Crear carpeta
    carpeta = os.path.join(os.getcwd(), "exportados", "Constancias de trabajo")
    ok, msg = crear_carpeta_segura(carpeta)
    if not ok:
        raise IOError(msg)

    nombre_base = sanitizar_nombre_archivo(f"ConstanciaTrabajo_{empleado['Cédula']}")
    nombre_archivo = os.path.join(carpeta, f"{nombre_base}.pdf")

    try:
        doc = SimpleDocTemplate(
            nombre_archivo,
            pagesize=letter,
            leftMargin=80,
            rightMargin=80,
            topMargin=180,
            bottomMargin=50
        )

        story = []

        # Título
        story.append(Paragraph("CONSTANCIA DE TRABAJO", styles["Title"]))
        story.append(Spacer(1, 16))

        # Texto principal
        director_ci = normalizar_cedula(institucion['director_ci'])
        texto = (
            f"Quien suscribe, {institucion['director']} Cédula de identidad {director_ci} "
            f"Director(a) de la {institucion['nombre']} hace constar "
            f"por medio de la presente que la ciudadana {empleado['Nombres']} {empleado['Apellidos']}, Cédula de Identidad "
            f"{cedula_normalizada} presta su servicio como {empleado['Cargo']} en esta institución desde "
            f"el {fecha_ingreso_str} hasta la presente fecha."
        )
        story.append(Paragraph(texto, justificado))
        story.append(Spacer(1, 40))

        # Fecha
        fecha_hoy = date.today().strftime("%d/%m/%Y")
        texto_fecha = (
            "Constancia que se expide a petición de la parte interesada en la ciudad de Puerto La Cruz, "
            f"a la fecha {fecha_hoy}."
        )
        story.append(Paragraph(texto_fecha, justificado))
        story.append(Spacer(1, 100))

        # Firma
        firma = f"________________________<br/>Prof. {institucion['director']}"
        story.append(Paragraph(firma, centrado))
        story.append(Spacer(1, 3))
        story.append(Paragraph(f"C.I. {director_ci}", centrado))
        story.append(Spacer(1, 3))
        story.append(Paragraph("Director", centrado))

        # Construir PDF
        doc.build(story, onFirstPage=encabezado_y_pie, onLaterPages=encabezado_y_pie)
        return nombre_archivo
        
    except Exception as e:
        raise IOError(f"Error generando PDF: {e}")


# REPORTES Y EXPORTACIONES

def exportar_reporte_pdf(parent, figure, titulo, criterio, etiquetas, valores, total) -> str:
    """Exporta un reporte estadístico a PDF con gráfica y tabla."""
    try:
        sugerido = f"reporte_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        ruta, _ = QFileDialog.getSaveFileName(
            parent,
            "Guardar reporte",
            sugerido,
            "PDF Files (*.pdf)"
        )
        
        if not ruta:
            return None
            
        if not ruta.endswith(".pdf"):
            ruta += ".pdf"

        # Guardar la figura en memoria
        buf = io.BytesIO()
        figure.savefig(buf, format="png", bbox_inches="tight", dpi=150)
        buf.seek(0)

        # Documento
        doc = SimpleDocTemplate(
            ruta,
            pagesize=letter,
            leftMargin=70,
            rightMargin=70,
            topMargin=180,
            bottomMargin=60
        )
        
        story = []

        # Título y fecha
        fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
        titulo_style = ParagraphStyle(
            'TituloReporte',
            parent=styles['Title'],
            fontSize=16,
            textColor=colors.HexColor("#2C3E50"),
            spaceAfter=8
        )
        story.append(Paragraph(f"<b>{titulo}</b>", titulo_style))
        story.append(Spacer(1, 5))
        
        # Subtítulo con fecha
        subtitulo_style = ParagraphStyle(
            'Subtitulo',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor("#7F8C8D"),
            alignment=TA_CENTER
        )
        story.append(Paragraph(f"Generado el {fecha}", subtitulo_style))
        story.append(Spacer(1, 20))

        # Calcular estadísticas
        if valores:
            promedio = sum(valores) / len(valores)
            maximo = max(valores)
            minimo = min(valores)
            
            # Encontrar categoría con valor máximo
            idx_max = valores.index(maximo)
            categoria_max = etiquetas[idx_max]
            
            # Calcular porcentaje del máximo
            porcentaje_max = (maximo / total * 100) if total > 0 else 0
        else:
            promedio = maximo = minimo = 0
            categoria_max = "N/A"
            porcentaje_max = 0

        # Crear tabla de resumen
        resumen_data = [
            ["RESUMEN", ""],
            ["Total de registros", f"{total}"],
            ["Promedio por categoría", f"{promedio:.1f}"],
            ["Valor máximo", f"{maximo} ({categoria_max})"],
            ["Porcentaje máximo", f"{porcentaje_max:.1f}%"],
            ["Categorías analizadas", f"{len(etiquetas)}"]
        ]
        
        resumen_table = Table(resumen_data, colWidths=[180, 180])
        resumen_table.setStyle(TableStyle([
            # Encabezado
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#3498DB")),
            ("TEXTCOLOR", (0,0), (-1,0), colors.whitesmoke),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,0), 11),
            ("ALIGN", (0,0), (-1,0), "CENTER"),
            ("SPAN", (0,0), (-1,0)),
            
            # Contenido
            ("BACKGROUND", (0,1), (0,-1), colors.HexColor("#ECF0F1")),
            ("BACKGROUND", (1,1), (1,-1), colors.white),
            ("FONTNAME", (0,1), (0,-1), "Helvetica-Bold"),
            ("FONTNAME", (1,1), (1,-1), "Helvetica"),
            ("FONTSIZE", (0,1), (-1,-1), 9),
            ("ALIGN", (0,1), (0,-1), "LEFT"),
            ("ALIGN", (1,1), (1,-1), "RIGHT"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            
            # Bordes
            ("BOX", (0,0), (-1,-1), 1.5, colors.HexColor("#3498DB")),
            ("LINEBELOW", (0,0), (-1,0), 1.5, colors.HexColor("#2980B9")),
            ("INNERGRID", (0,1), (-1,-1), 0.5, colors.HexColor("#BDC3C7")),
            
            # Padding
            ("LEFTPADDING", (0,0), (-1,-1), 10),
            ("RIGHTPADDING", (0,0), (-1,-1), 10),
            ("TOPPADDING", (0,0), (-1,-1), 6),
            ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ]))
        story.append(resumen_table)
        story.append(Spacer(1, 25))

        # Preparar datos con porcentajes
        data = [["Categoría", "Cantidad", "Porcentaje"]]
        for e, v in zip(etiquetas, valores):
            porcentaje = (v / total * 100) if total > 0 else 0
            data.append([str(e), str(v), f"{porcentaje:.1f}%"])
        
        # Fila de total
        data.append(["TOTAL", str(total), "100%"])

        # Crear tabla
        table = Table(data, hAlign="CENTER", colWidths=[190, 90, 80])
        
        # Estilo degradado para filas alternas
        table_style = [
            # Encabezado
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#34495E")),
            ("TEXTCOLOR", (0,0), (-1,0), colors.whitesmoke),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,0), 10),
            ("ALIGN", (0,0), (-1,-1), "CENTER"),
            
            # Contenido
            ("FONTNAME", (0,1), (-1,-2), "Helvetica"),
            ("FONTSIZE", (0,1), (-1,-2), 9),
            
            # Fila total
            ("BACKGROUND", (0,-1), (-1,-1), colors.HexColor("#E8F6F3")),
            ("TEXTCOLOR", (0,-1), (-1,-1), colors.HexColor("#117A65")),
            ("FONTNAME", (0,-1), (-1,-1), "Helvetica-Bold"),
            ("FONTSIZE", (0,-1), (-1,-1), 10),
            
            # Bordes
            ("BOX", (0,0), (-1,-1), 1.5, colors.HexColor("#34495E")),
            ("LINEBELOW", (0,0), (-1,0), 1.5, colors.HexColor("#2C3E50")),
            ("INNERGRID", (0,0), (-1,-1), 0.5, colors.HexColor("#BDC3C7")),
            
            # Padding
            ("LEFTPADDING", (0,0), (-1,-1), 8),
            ("RIGHTPADDING", (0,0), (-1,-1), 8),
            ("TOPPADDING", (0,0), (-1,-1), 5),
            ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ]
        
        # Agregar colores alternados para filas (excepto encabezado y total)
        for i in range(1, len(data) - 1):
            if i % 2 == 0:
                table_style.append(("BACKGROUND", (0,i), (-1,i), colors.HexColor("#F8F9F9")))
            else:
                table_style.append(("BACKGROUND", (0,i), (-1,i), colors.white))
        
        table.setStyle(TableStyle(table_style))
        story.append(table)
        story.append(Spacer(1, 25))

        story.append(Paragraph("<b>Representación Gráfica</b>", styles["Heading2"]))
        story.append(Spacer(1, 10))
        story.append(Image(buf, width=450, height=330))
        story.append(Spacer(1, 15))

        # ============== INTERPRETACIÓN AUTOMÁTICA ==============
        interpretacion_style = ParagraphStyle(
            'Interpretacion',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor("#34495E"),
            alignment=TA_JUSTIFY,
            leading=12,
            leftIndent=20,
            rightIndent=20,
            spaceBefore=5,
            spaceAfter=5
        )
        
        interpretacion = (
            f"<b>Análisis:</b> Se analizaron <b>{len(etiquetas)}</b> categorías según el criterio "
            f"<i>{criterio}</i>, con un total de <b>{total}</b> registros. "
            f"La categoría con mayor representación es <b>{categoria_max}</b> con <b>{maximo}</b> "
            f"registros ({porcentaje_max:.1f}% del total)."
        )
        
        if len(valores) > 1:
            # Calcular desviación entre máximo y mínimo
            rango = maximo - minimo
            if total > 0:
                variabilidad = (rango / total * 100)
                if variabilidad > 50:
                    interpretacion += " Se observa una <b>alta variabilidad</b> en la distribución."
                elif variabilidad > 20:
                    interpretacion += " La distribución muestra <b>variabilidad moderada</b>."
                else:
                    interpretacion += " La distribución es relativamente <b>homogénea</b>."
        
        story.append(Paragraph(interpretacion, interpretacion_style))
        story.append(Spacer(1, 20))

        # ============== OBSERVACIONES ==============
        story.append(Paragraph("<b>Observaciones Adicionales:</b>", styles["Normal"]))
        story.append(Spacer(1, 8))
        for _ in range(3):
            story.append(Paragraph("_" * 90, styles["Normal"]))
            story.append(Spacer(1, 10))

        # ============== FOOTER PERSONALIZADO ==============
        def footer_metadata(canvas, doc):
            """Footer con metadata del sistema"""
            canvas.saveState()
            
            # Primero dibujamos el encabezado y pie institucional
            encabezado(canvas, doc)
            pie_pagina(canvas, doc)
            
            # Línea decorativa separadora (debajo del pie institucional)
            canvas.setStrokeColor(colors.HexColor("#BDC3C7"))
            canvas.setLineWidth(0.5)
            canvas.line(70, 25, page_width - 70, 25)
            
            # Metadata del sistema (debajo de la línea)
            canvas.setFont("Helvetica", 7)
            canvas.setFillColor(colors.HexColor("#7F8C8D"))
            
            # Obtener usuario actual si está disponible
            usuario_texto = "Usuario: "
            if parent and hasattr(parent, 'usuario_actual'):
                usuario_texto += parent.usuario_actual.get('username', 'Sistema')
            else:
                usuario_texto += "Sistema"
            
            # Izquierda: Generado por SIRA
            canvas.drawString(70, 17, "Generado por SIRA v1.0")
            
            # Centro: Usuario
            text_width = canvas.stringWidth(usuario_texto, "Helvetica", 7)
            canvas.drawString((page_width - text_width) / 2, 17, usuario_texto)
            
            # Derecha: Número de página
            canvas.drawRightString(page_width - 70, 17, f"Página {doc.page}")
            
            canvas.restoreState()

        # Construir PDF con footer personalizado
        doc.build(story, onFirstPage=footer_metadata, onLaterPages=footer_metadata)
        buf.close()
        
        return ruta
        
    except Exception as e:
        if parent:
            crear_msgbox(
                parent,
                "Error",
                f"No se pudo exportar el reporte: {e}",
                QMessageBox.Critical
            ).exec()
        return None


def exportar_tabla_excel(nombre_archivo: str, encabezados: list, filas: list) -> str:
    """Exporta datos tabulares a un archivo Excel."""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos"

        # Escribir encabezados
        ws.append(encabezados)

        # Escribir filas (convertir objetos date/datetime a strings)
        for fila in filas:
            fila_procesada = []
            for celda in fila:
                if isinstance(celda, (date, datetime)):
                    fila_procesada.append(celda.strftime("%d/%m/%Y"))
                elif celda is None:
                    fila_procesada.append("")
                else:
                    fila_procesada.append(str(celda))
            ws.append(fila_procesada)

        # Guardar archivo
        wb.save(nombre_archivo)
        return nombre_archivo
        
    except Exception as e:
        raise IOError(f"Error generando archivo Excel: {e}")


def exportar_estudiantes_excel(parent, estudiantes: list) -> str:
    """Exporta lista de estudiantes a Excel."""
    if not estudiantes:
        if parent:
            crear_msgbox(
                parent,
                "Sin datos",
                "No hay estudiantes para exportar.",
                QMessageBox.Warning
            ).exec()
        return None

    try:
        sugerido = f"estudiantes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        ruta, _ = QFileDialog.getSaveFileName(
            parent,
            "Guardar reporte de estudiantes",
            sugerido,
            "Archivos Excel (*.xlsx)"
        )
        
        if not ruta:
            return None
            
        if not ruta.endswith(".xlsx"):
            ruta += ".xlsx"

        encabezados = list(estudiantes[0].keys())
        filas = [list(e.values()) for e in estudiantes]

        return exportar_tabla_excel(ruta, encabezados, filas)
        
    except Exception as e:
        if parent:
            crear_msgbox(
                parent,
                "Error",
                f"No se pudo exportar a Excel: {e}",
                QMessageBox.Critical
            ).exec()
        return None


def exportar_empleados_excel(parent, empleados: list) -> str:
    """Exporta lista de empleados a Excel."""
    if not empleados:
        if parent:
            crear_msgbox(
                parent,
                "Sin datos",
                "No hay empleados para exportar.",
                QMessageBox.Warning
            ).exec()
        return None

    try:
        sugerido = f"empleados_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        ruta, _ = QFileDialog.getSaveFileName(
            parent,
            "Guardar reporte de empleados",
            sugerido,
            "Archivos Excel (*.xlsx)"
        )
        
        if not ruta:
            return None
            
        if not ruta.endswith(".xlsx"):
            ruta += ".xlsx"

        encabezados = list(empleados[0].keys())
        filas = [list(e.values()) for e in empleados]

        return exportar_tabla_excel(ruta, encabezados, filas)
        
    except Exception as e:
        if parent:
            crear_msgbox(
                parent,
                "Error",
                f"No se pudo exportar a Excel: {e}",
                QMessageBox.Critical
            ).exec()
        return None


def generar_certificado_promocion_sexto(estudiante: dict, institucion: dict, año_escolar_egreso: str) -> str:
    """Genera certificado de promoción de 6to grado a 1er año de secundaria."""
    # Normalizar nombres de campos 
    estudiante_norm = {}
    
    # Mapear campos de BD a formato esperado
    estudiante_norm["Nombres"] = estudiante.get("Nombres") or estudiante.get("nombres", "")
    estudiante_norm["Apellidos"] = estudiante.get("Apellidos") or estudiante.get("apellidos", "")
    estudiante_norm["Cédula"] = estudiante.get("Cédula") or estudiante.get("cedula", "")
    estudiante_norm["Ciudad"] = estudiante.get("Ciudad") or estudiante.get("ciudad", "")
    estudiante_norm["Fecha Nac."] = estudiante.get("Fecha Nac.") or estudiante.get("fecha_nac_est")
    
    # Validar datos
    campos_est = ["Nombres", "Apellidos", "Cédula", "Ciudad", "Fecha Nac."]
    valido, mensaje = validar_datos_exportacion(estudiante_norm, campos_est)
    if not valido:
        raise ValueError(f"Datos de estudiante incompletos: {mensaje}")
    
    campos_inst = ["director", "director_ci", "nombre"]
    valido, mensaje = validar_datos_exportacion(institucion, campos_inst)
    if not valido:
        raise ValueError(f"Datos de institución incompletos: {mensaje}")
    
    if not año_escolar_egreso:
        raise ValueError("Año escolar de egreso no proporcionado")
    
    # Normalizar datos
    estudiante_norm["Nombres"] = str(estudiante_norm["Nombres"]).strip().upper()
    estudiante_norm["Apellidos"] = str(estudiante_norm["Apellidos"]).strip().upper()
    cedula_normalizada = normalizar_cedula(estudiante_norm["Cédula"], es_estudiante=True)
    
    # Convertir fecha
    fecha_nac_str = convertir_fecha_string(estudiante_norm['Fecha Nac.'])
    
    # Extraer año escolar (formato: "2023/2024" -> "2023-2024")
    año_periodo = str(año_escolar_egreso).replace('/', '-')
    
    # Crear carpeta
    carpeta = os.path.join(os.getcwd(), "exportados", "Certificados de promocion")
    ok, msg = crear_carpeta_segura(carpeta)
    if not ok:
        raise IOError(msg)

    nombre_base = sanitizar_nombre_archivo(f"Certificado_prosecusion_primaria{estudiante_norm['Cédula']}")
    nombre_archivo = os.path.join(carpeta, f"{nombre_base}.pdf")

    try:
        doc = SimpleDocTemplate(
            nombre_archivo,
            pagesize=letter,
            leftMargin=80,
            rightMargin=80,
            topMargin=180,
            bottomMargin=50
        )

        story = []

        # Título
        story.append(Paragraph("CERTIFICADO DE EDUCACIÓN PRIMARIA", styles["Title"]))
        story.append(Spacer(1, 16))

        # Texto principal
        director_ci = normalizar_cedula(institucion['director_ci'])
        # Intentar obtener ultima_seccion desde ambos formatos
        ultima_seccion = (estudiante.get('ultima_seccion') or 
                         estudiante.get('Ultima Seccion') or 
                         estudiante.get('letra') or 'N/A')
        
        texto = (
            f"Quien suscribe <b>{institucion['director'].upper()}</b> titular de la Cédula de Identidad "
            f"Nº <b>{director_ci}</b> Director(a) de la Institución Educativa "
            f"<b>{institucion['nombre'].upper()}</b>, ubicada en el Municipio <b>JUAN ANTONIO SOTILLO</b> de la Parroquia "
            f"<b>PUERTO LA CRUZ</b> adscrita al Centro de Desarrollo de la Calidad Educativa Estadal <b>ANZOÁTEGUI</b>. "
            f"Por la presente certifica que el (la) estudiante <b>{estudiante_norm['Apellidos']} {estudiante_norm['Nombres']}</b> "
            f"titular de Cédula Escolar Nº <b>{cedula_normalizada}</b>, nacido (a) en el Municipio "
            f"<b>{estudiante_norm['Ciudad']}</b> del Estado <b>ANZOÁTEGUI</b>, en fecha <b>{fecha_nac_str}</b>, cursó el "
            f"<b>6to Grado</b> correspondiéndole el literal <b>{ultima_seccion}</b> "
            f"durante el periodo escolar <b>{año_periodo}</b>, siendo promovido(a) al <b>1er Año del Nivel de "
            f"Educación Media</b>, previo cumplimiento a los requisitos establecidos en la normativa legal vigente."
        )
        story.append(Paragraph(texto, justificado))
        story.append(Spacer(1, 40))

        # Fecha de expedición
        fecha_hoy = date.today()
        dia = fecha_hoy.day
        mes_nombre = fecha_hoy.strftime("%B").upper()
        meses = {
            'JANUARY': 'ENERO', 'FEBRUARY': 'FEBRERO', 'MARCH': 'MARZO',
            'APRIL': 'ABRIL', 'MAY': 'MAYO', 'JUNE': 'JUNIO',
            'JULY': 'JULIO', 'AUGUST': 'AGOSTO', 'SEPTEMBER': 'SEPTIEMBRE',
            'OCTOBER': 'OCTUBRE', 'NOVEMBER': 'NOVIEMBRE', 'DECEMBER': 'DICIEMBRE'
        }
        mes_es = meses.get(mes_nombre, mes_nombre)
        año = fecha_hoy.year
        
        texto_fecha = f"Certificado que se expide en PUERTO LA CRUZ, a los {dia} días del mes de {mes_es} de {año}"
        story.append(Paragraph(texto_fecha, justificado))
        story.append(Spacer(1, 30))

        # Tabla de firmas institucionales
        tabla_firmas_data = [
            # Encabezados
            [
                Paragraph("<b>INSTITUCIÓN EDUCATIVA<br/>(PARA VALIDEZ NACIONAL)</b>", centrado),
                Paragraph("<b>CENTRO DE DESARROLLO DE LA CALIDAD<br/>EDUCATIVA ESTADAL<br/>(PARA VALIDEZ INTERNACIONAL)</b>", centrado)
            ],
            # Director(a)
            [
                Paragraph("DIRECTOR(A)", styles["Normal"]),
                Paragraph("DIRECTOR(A)", styles["Normal"])
            ],
            # Nombre y Apellido
            [
                Paragraph(f"Nombre y Apellido: {institucion['director'].upper()}", styles["Normal"]),
                Paragraph("Nombre y Apellido:", styles["Normal"])
            ],
            # Cédula de Identidad
            [
                Paragraph(f"Número de C.I: {director_ci}", styles["Normal"]),
                Paragraph("Número de C.I:", styles["Normal"])
            ],
            # Firma y Sello (espacios vacíos)
            [
                Paragraph("Firma y Sello:<br/><br/><br/>", styles["Normal"]),
                Paragraph("Firma y Sello:<br/><br/><br/>", styles["Normal"])
            ]
        ]

        tabla_firmas = Table(tabla_firmas_data, colWidths=[page_width/2 - 100, page_width/2 - 100])
        tabla_firmas.setStyle(TableStyle([
            # Bordes externos de la tabla
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            
            # Alineación
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),  # Encabezados centrados
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),   # Resto alineado a la izquierda
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),   # Alineación vertical superior
            
            # Padding interno
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        story.append(tabla_firmas)

        # Construir PDF con encabezado personalizado (sin pie de página)
        doc.build(story, onFirstPage=encabezado_prosecucion, onLaterPages=encabezado_prosecucion)
        return nombre_archivo
        
    except Exception as e:
        raise IOError(f"Error generando PDF: {e}")

def generar_constancia_retiro(estudiante: dict, institucion: dict, año_escolar: dict, motivo_retiro: str = None) -> str:
    """Genera constancia de retiro en PDF para un estudiante."""
    # Validar datos
    campos_est = ["Nombres", "Apellidos", "Cédula", "Grado", "Ciudad", "Fecha Nac."]
    valido, mensaje = validar_datos_exportacion(estudiante, campos_est)
    if not valido:
        raise ValueError(f"Datos de estudiante incompletos: {mensaje}")
    
    campos_inst = ["director", "director_ci", "nombre"]
    valido, mensaje = validar_datos_exportacion(institucion, campos_inst)
    if not valido:
        raise ValueError(f"Datos de institución incompletos: {mensaje}")
    
    # Extraer años
    año_inicio, año_fin = extraer_año_escolar(año_escolar)
    
    # Normalizar datos
    estudiante["Nombres"] = str(estudiante["Nombres"]).strip().upper()
    estudiante["Apellidos"] = str(estudiante["Apellidos"]).strip().upper()
    cedula_normalizada = normalizar_cedula(estudiante["Cédula"], es_estudiante=True)
    
    # Convertir fecha
    fecha_nac_str = convertir_fecha_string(estudiante['Fecha Nac.'])
    
    # Calcular edad
    try:
        if isinstance(estudiante['Fecha Nac.'], (date, datetime)):
            edad = calcular_edad(estudiante['Fecha Nac.'])
        else:
            fecha_obj = datetime.strptime(str(estudiante['Fecha Nac.']), "%d/%m/%Y").date()
            edad = calcular_edad(fecha_obj)
    except:
        edad = "N/A"
    
    # Motivo por defecto si no se proporciona
    if not motivo_retiro:
        motivo_retiro = "es retirado de la institución a solicitud de su representante siendo Promovido al siguiente grado"
    
    # Crear carpeta
    carpeta = os.path.join(os.getcwd(), "exportados", "Constancias de retiro")
    ok, msg = crear_carpeta_segura(carpeta)
    if not ok:
        raise IOError(msg)

    nombre_base = sanitizar_nombre_archivo(f"Constancia_retiro_{estudiante['Cédula']}")
    nombre_archivo = os.path.join(carpeta, f"{nombre_base}.pdf")

    try:
        doc = SimpleDocTemplate(
            nombre_archivo,
            pagesize=letter,
            leftMargin=80,
            rightMargin=80,
            topMargin=180,
            bottomMargin=50
        )

        story = []

        # Título
        story.append(Paragraph("CONSTANCIA DE RETIRO", styles["Title"]))
        story.append(Spacer(1, 20))

        # Texto principal
        director_ci = normalizar_cedula(institucion['director_ci'])
        
        # Determinar género para el artículo
        genero = estudiante.get("Género", "").lower()
        articulo = "el" if genero == "masculino" else "la"
        
        # Determinar si es nivel (Inicial) o grado (Primaria)
        tipo_educacion = estudiante.get("Tipo Educ.", "").lower()
        if "inicial" in tipo_educacion:
            grado_texto = f"{estudiante['Grado']}"
        else:
            grado_texto = f"{estudiante['Grado']} grado"
            
        texto = (
            f"La Dirección del plantel hace constar mediante la presente, que {articulo} Alumno(a): "
            f"<b>{estudiante['Apellidos']} {estudiante['Nombres']}</b>, "
            f"nacido(a) en <b>{estudiante['Ciudad'].upper()}</b> él <b>{fecha_nac_str}</b> "
            f"y de <b>{edad} año(s)</b> de edad, estudiante regular del <b>{grado_texto}</b> "
            f"para el año escolar <b>{año_inicio}-{año_fin}</b>, {motivo_retiro}.<br/><br/>"
        )
        story.append(Paragraph(texto, justificado))
        story.append(Spacer(1, 20))

        # Fecha actual
        fecha_hoy = date.today()
        dia = fecha_hoy.day
        mes_nombre = fecha_hoy.strftime("%B").upper()
        
        # Traducir mes al español
        meses = {
            'JANUARY': 'ENERO', 'FEBRUARY': 'FEBRERO', 'MARCH': 'MARZO',
            'APRIL': 'ABRIL', 'MAY': 'MAYO', 'JUNE': 'JUNIO',
            'JULY': 'JULIO', 'AUGUST': 'AGOSTO', 'SEPTEMBER': 'SEPTIEMBRE',
            'OCTOBER': 'OCTUBRE', 'NOVEMBER': 'NOVIEMBRE', 'DECEMBER': 'DICIEMBRE'
        }
        mes_es = meses.get(mes_nombre, mes_nombre)
        año = fecha_hoy.year
        
        texto_fecha = (
            f"Se expide constancia a solicitud de la parte interesada en Puerto La Cruz "
            f"a los <b>{dia}</b> días del mes de <b>{mes_es}</b> del año <b>{año}</b>."
        )
        story.append(Paragraph(texto_fecha, justificado))
        story.append(Spacer(1, 100))

        # Firma
        firma_texto = f"________________________<br/>Prof. {institucion['director']}"
        story.append(Paragraph(firma_texto, centrado))
        story.append(Spacer(1, 3))
        story.append(Paragraph(f"C.I. {director_ci}", centrado))
        story.append(Spacer(1, 3))
        story.append(Paragraph("Director", centrado))

        # Construir PDF
        doc.build(story, onFirstPage=encabezado_y_pie, onLaterPages=encabezado_y_pie)
        return nombre_archivo
        
    except Exception as e:
        raise IOError(f"Error generando PDF: {e}")


def generar_historial_estudiante_pdf(estudiante: dict, historial: list, institucion: dict) -> str:
    """Genera un PDF con el historial académico completo del estudiante."""
    # Validar datos
    campos_est = ["Nombres", "Apellidos", "Cédula"]
    valido, mensaje = validar_datos_exportacion(estudiante, campos_est)
    if not valido:
        raise ValueError(mensaje)
    
    # Normalizar datos
    estudiante["Nombres"] = str(estudiante["Nombres"]).strip().upper()
    estudiante["Apellidos"] = str(estudiante["Apellidos"]).strip().upper()
    
    # Normalizar cédula estudiantil (agregar CE- si no lo tiene)
    cedula = str(estudiante["Cédula"]).strip()
    if not cedula.upper().startswith("CE-"):
        cedula_estudiantil = f"CE-{cedula}"
    else:
        cedula_estudiantil = cedula.upper()
    
    # Crear carpeta
    carpeta = os.path.join(os.getcwd(), "exportados", "Historial academico")
    ok, msg = crear_carpeta_segura(carpeta)
    if not ok:
        raise IOError(msg)
    
    nombre_base = sanitizar_nombre_archivo(f"Historial_{estudiante['Cédula']}")
    nombre_archivo = os.path.join(carpeta, f"{nombre_base}.pdf")
    
    try:
        doc = SimpleDocTemplate(
            nombre_archivo,
            pagesize=letter,
            leftMargin=80,
            rightMargin=80,
            topMargin=180,
            bottomMargin=50
        )

        story = []
        
        # Título
        story.append(Paragraph("HISTORIAL ACADÉMICO", styles["Title"]))
        story.append(Spacer(1, 16))
        
        # Datos del estudiante
        datos_estudiante = f"""
        <b>Estudiante:</b> {estudiante['Nombres']} {estudiante['Apellidos']}<br/>
        <b>Cédula Estudiantil:</b> {cedula_estudiantil}<br/>
        """
        
        # Agregar grado actual si existe
        if "Grado" in estudiante and estudiante["Grado"]:
            datos_estudiante += f"<b>Grado actual:</b> {estudiante['Grado']}<br/>"
        
        story.append(Paragraph(datos_estudiante, styles["Normal"]))
        story.append(Spacer(1, 20))
        
        # Tabla de historial
        if historial:
            # Encabezados de la tabla
            datos_historial = [
                [
                    Paragraph("<b>Año Escolar</b>", centrado),
                    Paragraph("<b>Nivel</b>", centrado),
                    Paragraph("<b>Grado</b>", centrado),
                    Paragraph("<b>Sección</b>", centrado),
                    Paragraph("<b>Docente</b>", centrado)
                ]
            ]
            
            # Agregar filas del historial
            for registro in historial:
                año_escolar = f"{registro['año_inicio']}-{registro['año_inicio']+1}"
                datos_historial.append([
                    Paragraph(año_escolar, centrado),
                    Paragraph(str(registro['nivel']), centrado),
                    Paragraph(str(registro['grado']), centrado),
                    Paragraph(str(registro['letra']), centrado),
                    Paragraph(str(registro.get('docente', 'Sin asignar')), centrado)
                ])
            
            tabla_historial = Table(datos_historial, colWidths=[90, 110, 70, 70, 130])
            tabla_historial.setStyle(TableStyle([
                # Encabezado
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E5894')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                
                # Contenido
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                
                # Bordes y relleno
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F0F0')]),
                ('LEFTPADDING', (0, 0), (-1, -1), 5),
                ('RIGHTPADDING', (0, 0), (-1, -1), 5),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ]))
            
            story.append(tabla_historial)
        else:
            # Sin historial
            texto_sin_datos = Paragraph(
                "<i>No se encontró historial académico para este estudiante.</i>",
                justificado
            )
            story.append(texto_sin_datos)
        
        # Generar PDF
        doc.build(story, onFirstPage=encabezado_y_pie, onLaterPages=encabezado_y_pie)
        
        return nombre_archivo
        
    except Exception as e:
        raise IOError(f"Error generando historial en PDF: {e}")


def generar_reporte_rac(parent, empleados: list, institucion: dict) -> str:
    """Genera reporte RAC (Registro de Asignación de Cargos) en formato Excel."""
    if not empleados:
        crear_msgbox(
            parent,
            "Sin datos",
            "No hay empleados para exportar",
            QMessageBox.Warning
        ).exec()
        return None
    
    try:
        # Diálogo para guardar archivo
        fecha_actual = datetime.now().strftime("%Y%m%d")
        nombre_sugerido = f"Reporte_RAC_{fecha_actual}.xlsx"
        
        archivo, _ = QFileDialog.getSaveFileName(
            parent,
            "Guardar Reporte RAC",
            nombre_sugerido,
            "Archivos Excel (*.xlsx)"
        )
        
        if not archivo:
            return None
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "RAC"
        
        # Contar secciones activas del año escolar actual para especialistas
        total_secciones_activas = SeccionesModel.contar_activas_año_actual()
        
        # Datos fijos de la institución (valores predeterminados)
        DATOS_FIJOS = {
            'cod_estado': '2',
            'estado': 'ANZOATEGUI',
            'municipio': 'JUAN ANTONIO SOTILLO',
            'parroquia': 'PUERTO LA CRUZ',
            'codigo_dependencia': institucion.get('codigo_dependencia', '6562000'),
            'codigo_estadistico': institucion.get('codigo_estadistico', '31104'),
            'codigo_plantel': institucion.get('codigo_dea', 'OD03140321'),
            'nombre_plantel': institucion.get('nombre', 'E B DR SEVERIANO HERNANDEZ').upper(),
            'nivel': 'PRIMARIA',
            'modalidad': 'PRIMARIA',
            'ubicacion': 'URBANA',
            'turnos_plantel': 'INTEGRAL'
        }
        
        # Encabezados
        encabezados = [
            'COD ESTADO', 'ESTADO', 'MUNICIPIO', 'PARROQUIA',
            'CODIGO DEPENDENCIA', 'CODIGO ESTADISTICO', 'CODIGO DEL PLANTEL',
            'NOMBRE DEL PLANTEL EN NOMINA', 'NIVEL', 'MODALIDAD',
            'UBICACIÓN GEOGRAFICA', 'TURNOS QUE ATIENDE EL PLANTEL',
            'CODIGO RAC', 'CARGO', 'TIPO PERSONAL', 'CEDULA',
            'NOMBRE Y APELLIDO', 'FECHA DE INGRESO', 'SEXO',
            'HORAS ACADEMICAS', 'HORAS ADM', 'TURNO QUE ATIENDE',
            'GRADO QUE IMPARTE EL DOCENTE', 'SECCIÓN',
            'ESPECIALIDAD QUE IMPARTE EL DOCENTE', 'AÑO', 'SECCIONES',
            'MATERIA QUE IMPARTE O ESPECIALIDAD', 'PERIODO O GRUPO',
            'SITUACIÓN DEL TRABAJADOR', 'OBSERVACIÓN'
        ]
        
        ws.append(encabezados)
        
        # Función auxiliar para determinar tipo de personal
        def obtener_tipo_personal(cargo: str) -> str:
            """
            Determina el tipo de personal según el cargo.
            D = Docente, O = Obrero, A = Administrativo
            """
            cargo_upper = cargo.upper()
            
            # Docentes
            if any(x in cargo_upper for x in ['DOC', 'TSU EN EDUCACION']):
                return 'D'
            # Obreros
            elif 'OBRERO' in cargo_upper or 'COCINERA' in cargo_upper:
                return 'O'
            # Profesionales/Administrativos
            elif 'PROFESIONAL' in cargo_upper:
                return 'A'
            else:
                return 'O'  # Por defecto obrero
        
        # Función auxiliar para calcular horas según cargo
        def obtener_horas(cargo: str, tipo: str) -> tuple:
            """
            Retorna (horas_academicas, horas_adm) según el cargo.
            """
            cargo_upper = cargo.upper()
            
            # Director tiene horas administrativas
            if 'DIRECTOR' in cargo_upper:
                return '53 33', '53 33'
            
            # Docentes de aula tienen horas académicas
            if tipo == 'D' and 'AULA' in cargo_upper:
                return '53 33', '0'
            
            # Docentes generales
            if tipo == 'D':
                return '40', '0'
            
            # Resto (obreros, administrativos)
            return '40', '40'
        
        # Procesar cada empleado
        for empleado in empleados:
            # Normalizar cédula
            cedula = str(empleado.get('cedula', '')).strip()
            if cedula.startswith(('V-', 'E-', 'J-', 'G-')):
                cedula = cedula[2:]  # Remover prefijo
            cedula = cedula.replace('.', '').replace('-', '')
            
            # Normalizar nombres
            nombres = str(empleado.get('nombres', '')).strip().upper()
            apellidos = str(empleado.get('apellidos', '')).strip().upper()
            nombre_completo = f"{nombres} {apellidos}"
            
            # Fecha de ingreso
            fecha_ingreso = empleado.get('fecha_ingreso', '')
            if isinstance(fecha_ingreso, (date, datetime)):
                fecha_ingreso_str = fecha_ingreso.strftime('%d/%m/%Y')
            else:
                fecha_ingreso_str = str(fecha_ingreso)
            
            # Sexo
            genero = str(empleado.get('genero', '')).upper()
            sexo = 'M' if genero == 'MASCULINO' else 'F'
            
            # Cargo y tipo
            cargo = str(empleado.get('cargo', '')).strip().upper()
            tipo_personal_bd = empleado.get('tipo_personal', '')
            tipo_personal = tipo_personal_bd if tipo_personal_bd else obtener_tipo_personal(cargo)
            
            # Horas - usar valores reales de la BD, si no existen se usan valores calculados
            horas_acad_bd = empleado.get('horas_acad')
            horas_adm_bd = empleado.get('horas_adm')
            
            if horas_acad_bd is not None or horas_adm_bd is not None:
                # Usar valores de la BD, formatear con coma
                horas_acad = str(horas_acad_bd).replace('.', ',') if horas_acad_bd else '0'
                horas_adm = str(horas_adm_bd).replace('.', ',') if horas_adm_bd else '0'
            else:
                # Usar valores calculados por defecto
                horas_acad, horas_adm = obtener_horas(cargo, tipo_personal)
            
            # Código RAC
            codigo_rac = str(empleado.get('codigo_rac', '')).strip()
            
            # Estado del trabajador
            estado = empleado.get('estado', 'Activo')
            if isinstance(estado, int):
                situacion = 'ACTIVO' if estado == 1 else 'INACTIVO'
            else:
                situacion = str(estado).upper()
            
            # Determinar turno y grado según cargo
            turno = 'INTEGRAL'
            grado_imparte = ''
            seccion_imparte = ''
            
            # Verificar si es especialista
            especialidad = empleado.get('especialidad')
            
            if especialidad and especialidad.strip():
                # Es especialista
                grado_imparte = f"ESPECIALISTA EN {especialidad.upper()}"
                if total_secciones_activas > 0:
                    seccion_imparte = f"{total_secciones_activas} secciones"
                else:
                    seccion_imparte = "Todas"
            else:
                # No es especialista, verificar sección asignada
                seccion_grado = empleado.get('seccion_grado')
                seccion_letra = empleado.get('seccion_letra')
                seccion_nivel = empleado.get('seccion_nivel')
                
                if seccion_grado and seccion_letra and seccion_nivel:
                    grado_imparte = seccion_grado
                    # Si la letra es "Única", mostrar solo "U"
                    seccion_imparte = 'U' if seccion_letra.upper() == 'ÚNICA' else seccion_letra
                elif 'DIRECTOR' in cargo:
                    grado_imparte = 'DIRECTOR'
            
            # Construir fila
            fila = [
                DATOS_FIJOS['cod_estado'],
                DATOS_FIJOS['estado'],
                DATOS_FIJOS['municipio'],
                DATOS_FIJOS['parroquia'],
                DATOS_FIJOS['codigo_dependencia'],
                DATOS_FIJOS['codigo_estadistico'],
                DATOS_FIJOS['codigo_plantel'],
                DATOS_FIJOS['nombre_plantel'],
                DATOS_FIJOS['nivel'],
                DATOS_FIJOS['modalidad'],
                DATOS_FIJOS['ubicacion'],
                DATOS_FIJOS['turnos_plantel'],
                codigo_rac,
                cargo,
                tipo_personal,
                cedula,
                nombre_completo,
                fecha_ingreso_str,
                sexo,
                horas_acad,
                horas_adm,
                turno,
                grado_imparte,
                seccion_imparte,
                '',  # Especialidad
                '',  # Año
                '',  # Secciones
                '',  # Materia
                '',  # Periodo
                situacion,
                ''   # Observación
            ]
            
            ws.append(fila)
        
        # Aplicar estilos al reporte RAC
        # Colores y estilos
        azul_fondo_encabezado = PatternFill(start_color="4BACC6", end_color="4BACC6", fill_type="solid")
        azul_fondo_filas = PatternFill(start_color="B9CDE5", end_color="B9CDE5", fill_type="solid")
        blanco_texto = Font(name='Calibri', size=10, color="FFFFFF", bold=True)
        negro_texto = Font(name='Calibri', size=10, color="000000")
        borde_delgado = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        alineacion_centro = Alignment(horizontal='center', vertical='center', wrap_text=True)
        alineacion_izquierda = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Anchos de columna personalizados (en caracteres)
        anchos_columnas = [
            10,  # COD ESTADO
            15,  # ESTADO
            20,  # MUNICIPIO
            20,  # PARROQUIA
            18,  # CODIGO DEPENDENCIA
            18,  # CODIGO ESTADISTICO
            18,  # CODIGO DEL PLANTEL
            35,  # NOMBRE DEL PLANTEL
            12,  # NIVEL
            12,  # MODALIDAD
            15,  # UBICACIÓN
            15,  # TURNOS PLANTEL
            12,  # CODIGO RAC
            25,  # CARGO
            12,  # TIPO PERSONAL
            12,  # CEDULA
            35,  # NOMBRE Y APELLIDO
            15,  # FECHA INGRESO
            8,   # SEXO
            15,  # HORAS ACADEMICAS
            12,  # HORAS ADM
            15,  # TURNO
            25,  # GRADO IMPARTE
            12,  # SECCIÓN
            30,  # ESPECIALIDAD
            8,   # AÑO
            12,  # SECCIONES
            30,  # MATERIA
            15,  # PERIODO
            15,  # SITUACIÓN
            25   # OBSERVACIÓN
        ]
        
        # Aplicar anchos de columna
        for idx, ancho in enumerate(anchos_columnas, start=1):
            letra_columna = ws.cell(row=1, column=idx).column_letter
            ws.column_dimensions[letra_columna].width = ancho
        
        # Estilizar encabezado (fila 1)
        for col in range(1, len(encabezados) + 1):
            celda = ws.cell(row=1, column=col)
            celda.fill = azul_fondo_encabezado
            celda.font = blanco_texto
            celda.alignment = alineacion_centro
            celda.border = borde_delgado
        
        # Estilizar filas de datos
        for fila_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[fila_idx].height = 20  # Altura de fila
            for col in range(1, len(encabezados) + 1):
                celda = ws.cell(row=fila_idx, column=col)
                celda.fill = azul_fondo_filas
                celda.font = negro_texto
                celda.border = borde_delgado
                
                # Alineación según columna
                if col in [16, 17]:  # CEDULA, NOMBRE Y APELLIDO
                    celda.alignment = alineacion_izquierda
                else:
                    celda.alignment = alineacion_centro
        
        # Altura de la fila de encabezado
        ws.row_dimensions[1].height = 30
        
        # Guardar archivo
        wb.save(archivo)
        
        crear_msgbox(
            parent,
            "Exportación exitosa",
            f"Reporte RAC generado correctamente.\n\nTotal de empleados: {len(empleados)}\nArchivo: {os.path.basename(archivo)}",
            QMessageBox.Information
        ).exec()
        
        return archivo
        
    except Exception as e:
        crear_msgbox(
            parent,
            "Error",
            f"Error al generar reporte RAC:\n{str(e)}",
            QMessageBox.Critical
        ).exec()
        return None
